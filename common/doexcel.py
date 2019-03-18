import openpyxl
from openpyxl import load_workbook

class Cases:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.expected=None
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name=sheet_name
        self.wb = load_workbook(file_name)
    def get_data(self,sheet_name):
        sheet=self.wb[sheet_name]#获取sheet
        cases = []  # 所有的测试数据，都存在这个大列表中
        for i in range(2, sheet.max_row + 1):
            row_case = Cases()  # 实例化一个对象每一行的数据都存在子列表中，把每个值都存放在对象的属性里
            row_case.case_id = sheet.cell(i, 1).value  # 存的是第i行第一列的内容，也就是case_id
            row_case.title = sheet.cell(i, 2).value  # 存的是第i行第二列的内容，也就是title
            row_case.method= sheet.cell(i, 3).value  # 存的是第i行第二列的内容，也就是method
            row_case.url = sheet.cell(i, 4).value  # 存的是第i行第三列的内容，也就是接口url
            row_case.data = sheet.cell(i, 5).value  # 存的是第i行第四列的内容，也就是测试数据data
            row_case.expected = sheet.cell(i, 6).value  # 存的是第i行第五列的内容，也就是期望值expected
            cases.append(row_case)  # 把每组测试用例写回到大列表
        return cases

    def write_back(self,row,col,value,sheet_name):#写回测试结果
        # self.wb = load_workbook(self.file_name)
        sheet = self.wb[sheet_name]
        sheet.cell(row,col).value=value
        self.wb.save(self.file_name)
